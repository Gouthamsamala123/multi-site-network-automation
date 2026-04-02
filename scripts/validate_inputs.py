#!/usr/bin/env python3
"""
=============================================================================
MULTI-SITE DEPLOYMENT FRAMEWORK - INPUT VALIDATOR
Validates the Excel workbook for data integrity before config generation.
=============================================================================

Usage:
    python validate_inputs.py --input Framework.xlsx
"""

import sys
import re
import argparse
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] pip install openpyxl"); sys.exit(1)

VALID_DEVICE_TYPES = {"Switch", "Router", "SD-WAN", "Firewall", "WLC", "Access Point"}
VALID_STATUSES = {"Planning", "Ordered", "Received", "Configured", "Deployed", "Validated", "Production"}
HOSTNAME_PATTERN = re.compile(r'^[a-z]{2,3}-[a-z]{3}-[a-z]{2}-[a-z]{3}-\d+$')
IP_PATTERN = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
CIDR_PATTERN = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$')

errors = []
warnings = []

def error(sheet, row, msg):
    errors.append(f"[ERROR] {sheet} Row {row}: {msg}")

def warn(sheet, row, msg):
    warnings.append(f"[WARN]  {sheet} Row {row}: {msg}")

def ip_to_int(ip_str):
    parts = ip_str.split(".")
    return (int(parts[0]) << 24) + (int(parts[1]) << 16) + (int(parts[2]) << 8) + int(parts[3])

def validate_site_inventory(wb):
    ws = wb["Site Inventory"]
    hostnames = set()
    mgmt_ips = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_col=15, values_only=True), start=5):
        if not row[1]: continue
        site, loc, hostname, dtype = row[1], row[2], row[3], row[4]
        mgmt_ip, mgmt_vlan, mgmt_subnet = row[5], row[6], row[7]

        # Required fields
        if not hostname: error("Site Inventory", row_idx, "Hostname is required")
        if not dtype: error("Site Inventory", row_idx, "Device Type is required")
        if not mgmt_ip: error("Site Inventory", row_idx, "Management IP is required")

        # Device type validation
        if dtype and dtype not in VALID_DEVICE_TYPES:
            error("Site Inventory", row_idx, f"Invalid device type '{dtype}'. Valid: {VALID_DEVICE_TYPES}")

        # Hostname format check
        if hostname:
            if hostname in hostnames:
                error("Site Inventory", row_idx, f"Duplicate hostname: {hostname}")
            hostnames.add(hostname)
            if not HOSTNAME_PATTERN.match(str(hostname)):
                warn("Site Inventory", row_idx, f"Hostname '{hostname}' doesn't match naming convention: {{type}}-{{site}}-{{country}}-{{role}}-{{number}}")

        # IP validation
        if mgmt_ip and not IP_PATTERN.match(str(mgmt_ip)):
            error("Site Inventory", row_idx, f"Invalid management IP format: {mgmt_ip}")
        if mgmt_ip:
            if str(mgmt_ip) in mgmt_ips:
                error("Site Inventory", row_idx, f"Duplicate management IP: {mgmt_ip}")
            mgmt_ips.add(str(mgmt_ip))

        # Status
        status = row[13]
        if status and status not in VALID_STATUSES:
            warn("Site Inventory", row_idx, f"Non-standard status: {status}")

    return hostnames

def validate_ip_addressing(wb):
    ws = wb["IP Addressing"]
    subnets = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_col=13, values_only=True), start=5):
        if not row[1]: continue
        site, cidr, vlan_id, subnet = row[1], row[2], row[3], row[5]

        if not vlan_id: error("IP Addressing", row_idx, "VLAN ID is required")
        if not subnet: error("IP Addressing", row_idx, "Subnet CIDR is required")

        if subnet and not CIDR_PATTERN.match(str(subnet)):
            error("IP Addressing", row_idx, f"Invalid subnet CIDR: {subnet}")

        # Check for overlapping subnets within same site
        if subnet and CIDR_PATTERN.match(str(subnet)):
            parts = str(subnet).split("/")
            net_int = ip_to_int(parts[0])
            prefix = int(parts[1])
            size = 2 ** (32 - prefix)
            for prev_site, prev_net, prev_size, prev_row, prev_sub in subnets:
                if prev_site == site:
                    if not (net_int + size <= prev_net or prev_net + prev_size <= net_int):
                        error("IP Addressing", row_idx, f"Subnet {subnet} overlaps with {prev_sub} (row {prev_row})")
            subnets.append((site, net_int, size, row_idx, subnet))

def validate_vlan_mapping(wb):
    ws = wb["VLAN Mapping"]
    vlan_ids = defaultdict(set)

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_col=10, values_only=True), start=5):
        if not row[2]: continue
        site, vlan_id, vlan_name = row[1], row[2], row[3]

        if int(vlan_id) < 1 or int(vlan_id) > 4094:
            error("VLAN Mapping", row_idx, f"VLAN ID {vlan_id} out of range (1-4094)")

        key = f"{site}:{vlan_id}"
        if key in vlan_ids:
            error("VLAN Mapping", row_idx, f"Duplicate VLAN ID {vlan_id} for site {site}")
        vlan_ids[key] = True

def validate_routing(wb, valid_hostnames):
    ws = wb["Routing"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_col=9, values_only=True), start=5):
        if not row[1]: continue
        hostname = row[2]
        if hostname and hostname not in valid_hostnames:
            warn("Routing", row_idx, f"Hostname '{hostname}' not found in Site Inventory")

        dest = row[3]
        if dest and not CIDR_PATTERN.match(str(dest)) and dest != "0.0.0.0/0":
            warn("Routing", row_idx, f"Destination '{dest}' doesn't look like a valid CIDR")

def main(input_file):
    print(f"\n{'='*60}")
    print(f"  MULTI-SITE DEPLOYMENT FRAMEWORK - INPUT VALIDATOR")
    print(f"  File: {input_file}")
    print(f"{'='*60}\n")

    wb = load_workbook(input_file, data_only=True)

    print("[1/4] Validating Site Inventory...")
    hostnames = validate_site_inventory(wb)

    print("[2/4] Validating IP Addressing Plan...")
    validate_ip_addressing(wb)

    print("[3/4] Validating VLAN Mapping...")
    validate_vlan_mapping(wb)

    print("[4/4] Validating Routing Table...")
    validate_routing(wb, hostnames)

    print(f"\n{'='*60}")
    if errors:
        print(f"  RESULT: FAILED - {len(errors)} errors, {len(warnings)} warnings")
        print(f"{'='*60}\n")
        for e in errors: print(f"  {e}")
        if warnings:
            print()
            for w in warnings: print(f"  {w}")
        return False
    elif warnings:
        print(f"  RESULT: PASSED WITH WARNINGS - {len(warnings)} warnings")
        print(f"{'='*60}\n")
        for w in warnings: print(f"  {w}")
        return True
    else:
        print(f"  RESULT: PASSED - All validations clean")
        print(f"{'='*60}\n")
        return True

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", default="Multi_Site_Deployment_Framework.xlsx")
    args = parser.parse_args()
    success = main(args.input)
    sys.exit(0 if success else 1)
