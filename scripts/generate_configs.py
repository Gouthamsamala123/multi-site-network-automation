#!/usr/bin/env python3
"""
=============================================================================
MULTI-SITE NETWORK CONFIGURATION GENERATOR
Envista Holdings - Standardized Deployment Framework
=============================================================================
Reads the Multi-Site Deployment Framework Excel workbook and generates
per-device configurations using Jinja2 templates.

Usage:
    python generate_configs.py --input Framework.xlsx --output ./output
    python generate_configs.py --input Framework.xlsx --site "Versah"
    python generate_configs.py --input Framework.xlsx --device "sw-ver-us-mdf-1"
"""

import os
import sys
import argparse
from datetime import datetime

try:
    from openpyxl import load_workbook
    from jinja2 import Environment, FileSystemLoader
except ImportError:
    print("[ERROR] Missing dependencies. Install: pip install openpyxl jinja2")
    sys.exit(1)


def read_site_inventory(wb):
    """Parse Site Inventory sheet into device dictionaries."""
    ws = wb["Site Inventory"]
    devices = []
    for row in ws.iter_rows(min_row=5, max_col=15, values_only=True):
        if not row[1]:  # skip empty rows
            continue
        devices.append({
            "site_name": row[1],
            "site_location": row[2],
            "hostname": row[3],
            "device_type": row[4],
            "mgmt_ip": row[5],
            "mgmt_vlan": str(row[6]) if row[6] else "",
            "mgmt_subnet": row[7],
            "gateway": row[8],
            "isp_circuit": row[9],
            "serial": row[10],
            "stack_priority": row[11],
            "ha_role": row[12],
            "status": row[13],
            "notes": row[14],
        })
    return devices


def read_vlans(wb, site_name):
    """Parse VLAN Mapping sheet for a given site."""
    ws = wb["VLAN Mapping"]
    vlans = []
    for row in ws.iter_rows(min_row=5, max_col=10, values_only=True):
        if not row[2]:
            continue
        if row[1] in ("ALL SITES", site_name):
            vlans.append({
                "id": int(row[2]),
                "name": str(row[3]).replace(f"{site_name[:3].upper()}-", ""),
                "description": row[4],
                "subnet_template": row[5],
                "type": row[6],
                "dhcp": row[7],
                "voice": row[8],
                "zone": row[9],
            })
    return vlans


def read_ip_plan(wb, site_name):
    """Parse IP Addressing sheet for SVIs."""
    ws = wb["IP Addressing"]
    svis = []
    for row in ws.iter_rows(min_row=5, max_col=13, values_only=True):
        if not row[1] or row[1] != site_name:
            continue
        svis.append({
            "vlan_id": int(row[3]),
            "vlan_name": row[4],
            "subnet": row[5],
            "ip": row[6],
            "first_usable": row[7],
            "last_usable": row[8],
            "prefix": str(row[9]).replace("/", ""),
            "mask": row[10],
            "usable_hosts": row[11],
            "description": f"{row[4]} - {row[12]}" if row[12] else row[4],
        })
    return svis


def read_routes(wb, site_name, hostname=None):
    """Parse Routing sheet."""
    ws = wb["Routing"]
    routes = []
    for row in ws.iter_rows(min_row=5, max_col=9, values_only=True):
        if not row[1] or row[1] != site_name:
            continue
        if hostname and row[2] != hostname:
            continue
        dest_parts = str(row[3]).split("/")
        dest_ip = dest_parts[0]
        prefix_len = int(dest_parts[1]) if len(dest_parts) > 1 else 0
        masks = {0: "0.0.0.0", 8: "255.0.0.0", 16: "255.255.0.0",
                 23: "255.255.254.0", 24: "255.255.255.0", 25: "255.255.255.128",
                 27: "255.255.255.224", 28: "255.255.255.240", 32: "255.255.255.255"}
        routes.append({
            "destination": dest_ip,
            "mask": masks.get(prefix_len, "255.255.255.0"),
            "next_hop": row[4],
            "protocol": row[5],
            "interface": row[7],
            "description": row[8],
        })
    return routes


def build_site_context(device, vlans, svis, routes, global_vars):
    """Build the full Jinja2 context for a device."""
    site_prefix = device["hostname"].split("-")[1].upper() if "-" in device["hostname"] else "SITE"

    # Determine DHCP relay (gateway of Transfer VLAN)
    transfer_gw = None
    for svi in svis:
        if svi["vlan_id"] == 3:
            transfer_gw = svi.get("ip")
            break

    # Build SVI list with DHCP relay for user-facing VLANs
    svi_list = []
    for svi in svis:
        s = dict(svi)
        if svi["vlan_id"] in [10, 18, 188] and transfer_gw:
            s["dhcp_relay"] = transfer_gw
        svi_list.append(s)

    # Add L2-only VLANs as SVIs with N/A
    l2_vlans = [v for v in vlans if v["type"] in ("Guest",) or "L2" in str(v.get("subnet_template", ""))]
    for lv in l2_vlans:
        if not any(s["vlan_id"] == lv["id"] for s in svi_list):
            svi_list.append({
                "vlan_id": lv["id"],
                "vlan_name": lv["name"],
                "ip": "N/A",
                "mask": "N/A",
                "prefix": "",
                "description": f"{lv['name']} - L2 Only",
            })

    # Extract management subnet info
    mgmt_subnet_cidr = device.get("mgmt_subnet", "")
    mgmt_net = mgmt_subnet_cidr.split("/")[0] if "/" in mgmt_subnet_cidr else ""
    mgmt_prefix = mgmt_subnet_cidr.split("/")[1] if "/" in mgmt_subnet_cidr else "27"

    context = {
        "hostname": device["hostname"],
        "site_name": device["site_name"],
        "site_location": device["site_location"],
        "site_prefix": site_prefix,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "mgmt_ip": device["mgmt_ip"],
        "mgmt_prefix": mgmt_prefix,
        "vlans": vlans,
        "svis": svi_list,
        "static_routes": routes,
        "trunk_vlans": ",".join(str(v["id"]) for v in vlans),
        "ha_role": device.get("ha_role", "Standalone"),
        # Uplinks (standard pattern)
        "uplinks": [
            {"interface": "GigabitEthernet1/0/1", "peer": f"sd-{site_prefix.lower()}-mdf-1", "peer_port": "GE2"},
            {"interface": "GigabitEthernet1/0/2", "peer": f"sd-{site_prefix.lower()}-mdf-1", "peer_port": "GE3"},
            {"interface": "GigabitEthernet2/0/1", "peer": f"sd-{site_prefix.lower()}-mdf-2", "peer_port": "GE2"},
            {"interface": "GigabitEthernet2/0/2", "peer": f"sd-{site_prefix.lower()}-mdf-2", "peer_port": "GE3"},
        ],
        "ap_interfaces": [
            {"interface": "GigabitEthernet1/0/23", "ap_name": "AP01"},
            {"interface": "GigabitEthernet1/0/24", "ap_name": "AP03"},
            {"interface": "GigabitEthernet2/0/23", "ap_name": "AP02"},
            {"interface": "GigabitEthernet2/0/24", "ap_name": "AP04"},
        ],
        "user_port_ranges": ["GigabitEthernet1/0/3-22", "GigabitEthernet2/0/3-22"],
        # Global secrets (from global_vars or defaults)
        "enable_secret": global_vars.get("enable_secret", "<ENABLE_SECRET>"),
        "admin_user": global_vars.get("admin_user", "admin"),
        "admin_password": global_vars.get("admin_password", "<ADMIN_PASSWORD>"),
        "tacacs_server": global_vars.get("tacacs_server", "<TACACS_SERVER_IP>"),
        "tacacs_key": global_vars.get("tacacs_key", "<TACACS_KEY>"),
        "snmp_ro": global_vars.get("snmp_ro", "<RO_COMMUNITY>"),
        "snmp_rw": global_vars.get("snmp_rw", "<RW_COMMUNITY>"),
        "solarwinds_ip": global_vars.get("solarwinds_ip", "<SOLARWINDS_IP>"),
        "syslog_server": global_vars.get("syslog_server", "<SYSLOG_SERVER>"),
        "ntp_servers": global_vars.get("ntp_servers", ["<NTP_SERVER_1>", "<NTP_SERVER_2>"]),
        "mgmt_subnet": mgmt_net,
        "mgmt_wildcard": "0.0.0.31",
    }

    if device.get("stack_priority"):
        context["stack_priority"] = device["stack_priority"]
        context["stack_member"] = "1"

    return context


def generate(input_file, output_dir, template_dir, site_filter=None, device_filter=None):
    """Main generation function."""
    print(f"\n{'='*60}")
    print(f"  MULTI-SITE CONFIG GENERATOR")
    print(f"  Input:  {input_file}")
    print(f"  Output: {output_dir}")
    print(f"  Time:   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    wb = load_workbook(input_file, data_only=True)
    env = Environment(loader=FileSystemLoader(template_dir), trim_blocks=True, lstrip_blocks=True)

    devices = read_site_inventory(wb)
    print(f"[INFO] Loaded {len(devices)} devices from Site Inventory")

    # Global vars (could come from a secrets file in production)
    global_vars = {}

    generated = 0
    skipped = 0

    for device in devices:
        if site_filter and device["site_name"] != site_filter:
            skipped += 1; continue
        if device_filter and device["hostname"] != device_filter:
            skipped += 1; continue

        dtype = device["device_type"]
        hostname = device["hostname"]
        site = device["site_name"]

        # Select template
        if dtype == "Switch":
            template_name = "switch_config.j2"
        elif dtype == "SD-WAN":
            template_name = "velocloud_edge.j2"
        else:
            print(f"  [SKIP] {hostname} ({dtype}) - no template for this device type")
            skipped += 1; continue

        vlans = read_vlans(wb, site)
        svis = read_ip_plan(wb, site)
        routes = read_routes(wb, site, hostname)
        context = build_site_context(device, vlans, svis, routes, global_vars)

        # VeloCloud Edge specific fields
        if dtype == "SD-WAN":
            context["edge_model"] = "edge640"
            context["serial_number"] = device.get("serial", "TBD")
            context["activation_key"] = global_vars.get("activation_key", "<ACTIVATION_KEY>")
            context["profile_id"] = global_vars.get("profile_id", "<SITE_PROFILE_ID>")
            context["software_version"] = "5.4.0"
            context["ha_mode"] = "active-standby"
            context["ha_peer_serial"] = "<PEER_SERIAL>"
            context["wan_interface"] = "GE1"
            context["ha_interface"] = "GE2"
            context["isp1_gw"] = "<ISP1_GATEWAY>"
            context["isp2_gw"] = "<ISP2_GATEWAY>"
            context["dhcp_relay"] = global_vars.get("dhcp_relay", "<DHCP_SERVER_IP>")
            context["dns_primary"] = global_vars.get("dns_primary", "<DNS_SERVER_1>")
            context["dns_secondary"] = global_vars.get("dns_secondary", "<DNS_SERVER_2>")
            context["internet_policy"] = "directInternetBreakout"
            # LAN interfaces for VeloCloud
            all_vlan_ids = [v["id"] for v in vlans]
            context["lan_interfaces"] = [
                {"name": "GE3", "type": "TRUNK", "description": f"LAN-TO-{context['site_prefix']}-SWITCH-1",
                 "speed": "1000", "duplex": "FULL", "vlans": all_vlan_ids},
                {"name": "GE4", "type": "TRUNK", "description": f"LAN-TO-{context['site_prefix']}-SWITCH-2",
                 "speed": "1000", "duplex": "FULL", "vlans": all_vlan_ids},
            ]

        try:
            template = env.get_template(template_name)
            config = template.render(**context)
        except Exception as e:
            print(f"  [ERROR] {hostname}: {e}")
            continue

        # Write output
        site_dir = os.path.join(output_dir, site.replace(" ", "_"))
        os.makedirs(site_dir, exist_ok=True)
        ext = ".json" if dtype == "SD-WAN" else ".cfg"
        out_file = os.path.join(site_dir, f"{hostname}{ext}")
        with open(out_file, "w") as f:
            f.write(config)

        print(f"  [OK] {hostname}{ext} ({dtype}) -> {site_dir}/")
        generated += 1

    print(f"\n{'='*60}")
    print(f"  SUMMARY: {generated} configs generated, {skipped} skipped")
    print(f"  Output directory: {output_dir}")
    print(f"{'='*60}\n")

    return generated


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Multi-Site Network Config Generator")
    parser.add_argument("--input", "-i", default="Multi_Site_Deployment_Framework.xlsx", help="Excel input file")
    parser.add_argument("--output", "-o", default="./output", help="Output directory")
    parser.add_argument("--templates", "-t", default="./templates", help="Jinja2 template directory")
    parser.add_argument("--site", "-s", default=None, help="Generate for specific site only")
    parser.add_argument("--device", "-d", default=None, help="Generate for specific device only")
    args = parser.parse_args()

    generate(args.input, args.output, args.templates, args.site, args.device)
